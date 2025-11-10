# app/routers/export.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import os
from typing import Any, Dict, List, Optional, Tuple, Set

import requests
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.comments import Comment
import openpyxl.packaging.manifest as _manifest

router = APIRouter()

# -----------------------------------------------------------------------------
# Helpers: reuse your own API so filters match the Comparison tab 1:1
# -----------------------------------------------------------------------------
try:
    if ".webp" not in _manifest.mimetypes.types_map[True]:
        _manifest.mimetypes.add_type("image/webp", ".webp")
        _manifest.mimetypes.add_type("image/webp", ".WEBP")
except Exception:
    pass


def _internal_base() -> str:
    return os.getenv("EMAIL_INTERNAL_BASE_URL", f"http://127.0.0.1:{os.getenv('APP_PORT','8001')}")


def _fetch_compare(
    site_code: str,
    limit: int,
    source: str,
    q: Optional[str],
    tag_id: Optional[str],
    brand: Optional[str],
    price_filter: Optional[str],
    category_id: Optional[str],
    praktis_presence: Optional[str],
) -> List[Dict[str, Any]]:
    params: Dict[str, str] = {
        "site_code": site_code,
        "limit": str(limit),
        "source": source,
    }
    if q:                params["q"] = q
    if tag_id:           params["tag_id"] = str(tag_id)
    if brand:            params["brand"] = brand
    if price_filter:     params["price_filter"] = price_filter
    if category_id:      params["category_id"] = category_id
    if praktis_presence: params["praktis_presence"] = praktis_presence

    url = f"{_internal_base()}/api/compare"
    r = requests.get(url, params=params, timeout=300)
    r.raise_for_status()
    return r.json()


def _fetch_assets(skus: List[str]) -> Dict[str, Dict[str, Any]]:
    if not skus:
        return {}
    url = f"{_internal_base()}/api/products/assets"
    r = requests.get(url, params={"skus": ",".join(skus)}, timeout=120)
    if r.status_code != 200:
        return {}
    return r.json()


def _to_num(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s or s in ("n/a", "none"):
        return None
    try:
        return float(s.replace(" ", "").replace(",", "."))
    except Exception:
        return None


def _effective(promo: Optional[float], regular: Optional[float]) -> Optional[float]:
    return promo if promo is not None else regular


def _pivot_all(flat_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    product_sku -> single record with praktis price + best competitor price per site.
    For competitors we use **effective price** (promo if available, else regular) and
    carry the label + both promo/regular for display purposes.
    """
    out: Dict[str, Dict[str, Any]] = {}
    for r in flat_rows:
        sku = r.get("product_sku") or ""
        if not sku:
            continue

        g = out.setdefault(
            sku,
            {
                "product_sku": sku,
                "product_name": r.get("product_name"),
                "praktis_price": _to_num(r.get("product_price_regular")),
                # each competitor keeps: effective price, url, label, promo, regular
                "praktiker":   {"price": None, "url": None, "label": None, "promo": None, "regular": None},
                "mrbricolage": {"price": None, "url": None, "label": None, "promo": None, "regular": None},
                "mashinibg":   {"price": None, "url": None, "label": None, "promo": None, "regular": None},
            },
        )

        site = (r.get("competitor_site") or "").lower()
        reg = _to_num(r.get("competitor_price_regular"))
        prm = _to_num(r.get("competitor_price_promo"))
        eff = _effective(prm, reg)
        url = r.get("competitor_url")
        lbl = r.get("competitor_label")

        def maybe_update(bucket: Dict[str, Any]):
            cur = bucket["price"]
            if cur is None or (eff is not None and eff < cur):
                bucket["price"]   = eff
                bucket["url"]     = url
                bucket["label"]   = lbl
                bucket["promo"]   = prm
                bucket["regular"] = reg

        if "praktiker" in site:
            maybe_update(g["praktiker"])
        elif "bricol" in site:
            maybe_update(g["mrbricolage"])
        elif "mashin" in site:
            maybe_update(g["mashinibg"])
    return out


# -----------------------------------------------------------------------------
# Excel styling helpers
# -----------------------------------------------------------------------------
_thin = Side(style="thin", color="999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_hdr = PatternFill("solid", fgColor="F2F2F2")

_fill_green = PatternFill("solid", fgColor="DFF0D8")
_fill_red = PatternFill("solid", fgColor="F8D7DA")

_GRAY = "999999"


def _ws_header(ws, headers: List[str], widths_px: List[int]):
    row = 1 if (ws.max_row == 1 and ws.cell(row=1, column=1).value in (None, "")) else ws.max_row + 1
    for i, (h, wpx) in enumerate(zip(headers, widths_px), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=False)
        c.fill = _hdr
        c.border = _border
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wpx / 7.0


def _ws_row(ws, values: List[Any], number_cols: Set[int]):
    row = ws.max_row + 1
    for i, v in enumerate(values, start=1):
        if isinstance(v, dict) and v.get("__HYPERLINK__"):
            cell = ws.cell(row=row, column=i, value=v["text"])
            cell.hyperlink = v["url"]
        else:
            cell = ws.cell(row=row, column=i, value=v)
        cell.border = _border
        cell.alignment = Alignment(
            horizontal=("right" if i in number_cols else "left"),
            vertical="center",
            wrap_text=True,  # allow comments/info to show nicely
        )


def _set_fill(ws, row_idx: int, col_idx: int, fill: PatternFill | None):
    if fill:
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _fmt_money(v):
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return v


# -----------------------------------------------------------------------------
# Columns support (tokens sent from the frontend)
# -----------------------------------------------------------------------------
ALL_COLS: Dict[str, Tuple[str, int]] = {
    "praktis_code": ("Praktis Code", 110),
    "praktis_name": ("Praktis Name", 360),
    "praktis_price": ("Praktis Regular Price", 140),
    "praktiker_price": ("Praktiker Price", 170),
    "mrbricolage_price": ("MrBricolage Price", 190),
    "mashinibg_price": ("OnlineMashini Price", 200),
}

SITE_COLS: Dict[str, Tuple[str, int]] = {
    "praktis_code": ("Praktis Code", 110),
    "praktis_name": ("Praktis Name", 360),
    "competitor_code": ("Competitor Code", 160),
    "competitor_name": ("Competitor Name", 320),
    "praktis_regular": ("Praktis Regular Price", 140),
    "competitor_regular": ("Competitor Regular Price", 170),
    "praktis_promo": ("Praktis Promo Price", 150),
    "competitor_promo": ("Competitor Promo Price", 160),
    "competitor_label": ("Competitor Label", 150),
}


def _parse_columns(raw: Optional[str], site_code: str) -> List[str]:
    if raw:
        toks = [t.strip().lower() for t in raw.split(",") if t.strip()]
    else:
        toks = []
    allowed = ALL_COLS if site_code == "all" else SITE_COLS
    if toks:
        # drop any stray "image" tokens from old frontends
        toks = [t for t in toks if t != "image"]
        return [t for t in toks if t in allowed]
    if site_code == "all":
        return ["praktis_code", "praktis_name", "praktis_price",
                "praktiker_price", "mrbricolage_price", "mashinibg_price"]
    else:
        return ["praktis_code", "praktis_name",
                "competitor_code", "competitor_name",
                "praktis_regular", "competitor_regular",
                "praktis_promo", "competitor_promo"]


# -----------------------------------------------------------------------------
# Public endpoint
# -----------------------------------------------------------------------------
@router.get("/export/compare.xlsx")
def export_compare_xlsx(
    site_code: str,
    limit: int = Query(2000, ge=1, le=100000),
    source: str = "snapshots",
    q: Optional[str] = None,
    tag_id: Optional[str] = None,
    brand: Optional[str] = None,
    price_status: Optional[str] = Query(None, pattern="^(ours_lower|ours_higher)$"),
    page: Optional[int] = None,
    per_page: Optional[int] = None,
    columns: Optional[str] = Query(None, description="Comma-separated column ids in the current order"),
    category_id: Optional[str] = None,
    praktis_presence: Optional[str] = Query(None, pattern="^(present|missing)$"),
):
    """
    XLSX export that mirrors the Comparison tab (without image column).
    """
    limit_for_compare = min(int(limit), 2000)

    rows = _fetch_compare(
        site_code=site_code,
        limit=limit_for_compare,
        source=source,
        q=q,
        tag_id=tag_id,
        brand=brand,
        price_filter=price_status,
        category_id=category_id,
        praktis_presence=praktis_presence,
    )

    # columns + headers
    col_tokens = _parse_columns(columns, site_code)
    allowed = ALL_COLS if site_code == "all" else SITE_COLS
    headers = [allowed[t][0] for t in col_tokens]
    widths = [allowed[t][1] for t in col_tokens]

    # assets (only for hyperlinks to Praktis code/name; image is removed)
    need_assets = any(t in col_tokens for t in ("praktis_code", "praktis_name"))
    skus = [r.get("product_sku") for r in rows if r.get("product_sku")]
    assets = _fetch_assets(list(dict.fromkeys(skus))) if need_assets else {}

    # Praktis presence
    if praktis_presence in ("present", "missing"):
        has_url = {sku for sku, a in (assets or {}).items() if a and a.get("product_url")}
        if praktis_presence == "present":
            rows = [r for r in rows if r.get("product_sku") in has_url]
        else:
            rows = [r for r in rows if r.get("product_sku") not in has_url]

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    _ws_header(ws, headers, widths)

    number_cols: Set[int] = set()  # 1-based indices of numeric columns

    if site_code == "all":
        # Pivot after presence-filtering
        pivot = _pivot_all(rows)

        # Only SKUs that have price in any visible competitor (so we don't leak other stores)
        visible_comp_cols = [t for t in col_tokens if t in ("praktiker_price", "mrbricolage_price", "mashinibg_price")]

        def _has_visible_price(g: Dict[str, Any]) -> bool:
            if not visible_comp_cols:
                return True
            for tok in visible_comp_cols:
                key = "praktiker" if tok.startswith("praktiker") else ("mrbricolage" if tok.startswith("mrbricolage") else "mashinibg")
                if _to_num(g[key]["price"]) is not None:
                    return True
            return False

        # Filter and paginate here (after all filters)
        skus_filtered = [sku for sku, g in pivot.items() if _has_visible_price(g)]
        if page and per_page:
            start = max(0, (page - 1) * per_page)
            skus_filtered = skus_filtered[start: start + per_page]

        for sku in skus_filtered:
            g = pivot[sku]
            a = assets.get(sku) or {}
            praktis_url = a.get("product_url") if need_assets else None

            row_vals: List[Any] = []
            price_col_map: Dict[str, int] = {}

            for idx0, tok in enumerate(col_tokens, start=1):
                if tok == "praktis_code":
                    row_vals.append({"__HYPERLINK__": True, "text": sku, "url": praktis_url} if praktis_url else sku)
                elif tok == "praktis_name":
                    name = g.get("product_name") or "Няма име"
                    row_vals.append({"__HYPERLINK__": True, "text": name, "url": praktis_url} if praktis_url else name)
                elif tok == "praktis_price":
                    val = _fmt_money(g.get("praktis_price"))
                    row_vals.append(val)
                    price_col_map["praktis_price"] = idx0
                    number_cols.add(idx0)
                elif tok in ("praktiker_price", "mrbricolage_price", "mashinibg_price"):
                    key = "praktiker" if tok.startswith("praktiker") else ("mrbricolage" if tok.startswith("mrbricolage") else "mashinibg")
                    eff = g[key]["price"]
                    url = g[key]["url"]
                    label = g[key]["label"]
                    prm = g[key]["promo"]
                    reg = g[key]["regular"]

                    # Display like UI: promo shown with badge "П" if promo exists; keep link
                    display_text = None
                    if prm is not None and (eff == prm):
                        badge = " П"
                        display_text = f"{prm:.2f}{badge}"
                    else:
                        # no promo → show effective/regular number
                        if eff is not None:
                            display_text = f"{eff:.2f}"

                    if url and display_text is not None:
                        row_vals.append({"__HYPERLINK__": True, "text": display_text, "url": url})
                    else:
                        row_vals.append(display_text if display_text is not None else None)

                    # Add comment with regular price or label details if useful
                    price_col_map[tok] = idx0
                    number_cols.add(idx0)
                else:
                    row_vals.append(None)

            _ws_row(ws, row_vals, number_cols)
            row_idx = ws.max_row

            # Add comments for competitor prices where promo exists to show regular/label
            for tok in ("praktiker_price", "mrbricolage_price", "mashinibg_price"):
                if tok not in price_col_map:
                    continue
                key = "praktiker" if tok.startswith("praktiker") else ("mrbricolage" if tok.startswith("mrbricolage") else "mashinibg")
                prm = g[key]["promo"]; reg = g[key]["regular"]; lbl = g[key]["label"]
                if prm is not None and reg is not None:
                    c = ws.cell(row=row_idx, column=price_col_map[tok])
                    try:
                        msg = f"Regular: {reg:.2f}"
                        if lbl: msg += f" | Label: {lbl}"
                        c.comment = Comment(msg, "export")
                    except Exception:
                        pass

            # Highlight: min among visible numeric price columns only (praktis + chosen comps)
            visible_cols_for_min = ["praktis_price"] + visible_comp_cols
            visible_nums: List[Tuple[str, float, int]] = []
            for tok in visible_cols_for_min:
                if tok not in price_col_map and tok != "praktis_price":
                    continue
                if tok == "praktis_price":
                    v = _to_num(g.get("praktis_price"))
                    if v is not None:
                        visible_nums.append((tok, v, price_col_map.get(tok, col_tokens.index("praktis_price")+1)))
                else:
                    key = "praktiker" if tok.startswith("praktiker") else ("mrbricolage" if tok.startswith("mrbricolage") else "mashinibg")
                    v = _to_num(g[key]["price"])
                    if v is not None:
                        visible_nums.append((tok, v, price_col_map[tok]))

            if visible_nums:
                mn = min(v for _, v, _ in visible_nums)
                for _, v, col in visible_nums:
                    if v == mn:
                        _set_fill(ws, row_idx, col, _fill_green)
                    elif v > mn:
                        _set_fill(ws, row_idx, col, _fill_red)

    else:
        # Single-site view
        if page and per_page:
            start = max(0, (page - 1) * per_page)
            rows = rows[start: start + per_page]

        for r in rows:
            sku = r.get("product_sku") or ""
            a = assets.get(sku) or {}
            praktis_url = a.get("product_url") if need_assets else None

            comp_name = r.get("competitor_name") or "N/A"
            comp_url = r.get("competitor_url")
            our_reg = _to_num(r.get("product_price_regular"))
            comp_reg = _to_num(r.get("competitor_price_regular"))
            our_prm = _to_num(r.get("product_price_promo"))
            comp_prm = _to_num(r.get("competitor_price_promo"))
            comp_code = r.get("competitor_sku")
            comp_lbl = r.get("competitor_label")

            row_vals: List[Any] = []
            praktis_reg_col, comp_reg_col, comp_prm_col = None, None, None

            for idx0, tok in enumerate(col_tokens, start=1):
                if tok == "praktis_code":
                    row_vals.append({"__HYPERLINK__": True, "text": sku, "url": praktis_url} if praktis_url else sku)
                elif tok == "praktis_name":
                    name = r.get("product_name") or "Няма име"
                    row_vals.append({"__HYPERLINK__": True, "text": name, "url": praktis_url} if praktis_url else name)
                elif tok == "competitor_code":
                    row_vals.append(comp_code)
                elif tok == "competitor_name":
                    row_vals.append({"__HYPERLINK__": True, "text": comp_name, "url": comp_url} if comp_url else comp_name)
                elif tok == "praktis_regular":
                    row_vals.append(_fmt_money(our_reg))
                    praktis_reg_col = idx0
                    number_cols.add(idx0)
                elif tok == "competitor_regular":
                    row_vals.append(_fmt_money(comp_reg))
                    comp_reg_col = idx0
                    number_cols.add(idx0)
                elif tok == "praktis_promo":
                    row_vals.append(_fmt_money(our_prm))
                    number_cols.add(idx0)
                elif tok == "competitor_promo":
                    if comp_prm is not None:
                        badge = " П"
                        text = f"{comp_prm:.2f}{badge}" if comp_lbl else f"{comp_prm:.2f}"
                        row_vals.append(text)
                    else:
                        row_vals.append(_fmt_money(comp_prm))
                    comp_prm_col = idx0
                    number_cols.add(idx0)
                elif tok == "competitor_label":
                    row_vals.append(comp_lbl)
                else:
                    row_vals.append(None)

            _ws_row(ws, row_vals, number_cols)
            row_idx = ws.max_row

            # Strike-through the competitor regular price if there is a promo
            if comp_prm is not None and comp_reg is not None and comp_reg_col:
                c = ws.cell(row=row_idx, column=comp_reg_col)
                c.font = Font(color=_GRAY, strike=True)

            # Highlight only if both regular price columns are visible
            if praktis_reg_col and comp_reg_col and (our_reg is not None) and (comp_reg is not None):
                if our_reg < comp_reg:
                    _set_fill(ws, row_idx, praktis_reg_col, _fill_green)
                    _set_fill(ws, row_idx, comp_reg_col, _fill_red)
                elif our_reg > comp_reg:
                    _set_fill(ws, row_idx, praktis_reg_col, _fill_red)
                    _set_fill(ws, row_idx, comp_reg_col, _fill_green)

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"comparison_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'}
    )
