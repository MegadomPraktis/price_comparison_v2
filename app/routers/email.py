# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import os
import ssl
import smtplib
import asyncio
import requests
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any, Tuple

from fastapi import APIRouter, HTTPException, Body, Query
from fastapi.responses import JSONResponse
from sqlalchemy import select, func  # ← added func

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.db import get_session
from app.models import (
    EmailRule, EmailWeeklySchedule, PriceSubset,
    Product, Group, CompetitorSite, PriceSnapshot  # ← added CompetitorSite, PriceSnapshot
)
from app.schemas import EmailRuleIn, EmailRuleOut, WeeklySchedule

router = APIRouter()

# ========================== SMTP / runtime settings ===========================
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_SECURITY = os.getenv("SMTP_SECURITY", "starttls").lower()  # starttls|ssl|plain
SMTP_TLS_FLAG = os.getenv("SMTP_TLS", None)   # "1"/"0"
SMTP_SSL_FLAG = os.getenv("SMTP_SSL", None)   # "1"/"0"
SMTP_DEBUG = os.getenv("SMTP_DEBUG", "0").lower() in ("1","true","yes","on")
SMTP_TIMEOUT = float(os.getenv("SMTP_TIMEOUT", "20"))
SENDER = os.getenv("SMTP_SENDER", SMTP_USER or "no-reply@example.com")
OUTBOX_DIR = os.getenv("SMTP_OUTBOX_DIR", os.path.join(os.path.dirname(__file__), ".", ".", "outbox"))
APP_BASE = os.getenv("EMAIL_INTERNAL_BASE_URL", f"http://127.0.0.1:{os.getenv('APP_PORT', '8001')}")

# ================================ XLSX styles ================================
_thin = Side(style="thin", color="999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_head = PatternFill("solid", fgColor="0F1B38")
_center = Alignment(horizontal="center", vertical="center")
_right = Alignment(horizontal="right", vertical="center")
_mid = Alignment(vertical="center")
_green = PatternFill("solid", fgColor="E9F6E9")
_red   = PatternFill("solid", fgColor="FBE4E6")

def _px_to_w(px: int) -> float:
    return max(8.0, px / 7.0)

# ============================ Internal API calls =============================
def _get_json(url: str, params: dict) -> Any:
    r = requests.get(url, params=params, timeout=600)
    r.raise_for_status()
    return r.json()

def _fetch_compare(site_code: str, tag_id: Optional[int], brand: Optional[str], q: Optional[str] = None) -> List[Dict[str, Any]]:
    params = {"site_code": site_code, "limit": "2000", "source": "snapshots"}
    if tag_id not in (None, "", "all"):
        params["tag_id"] = str(tag_id)
    if brand:
        params["brand"] = brand
    if q:
        params["q"] = q
    return _get_json(f"{APP_BASE}/api/compare", params)

def _fetch_assets(skus: List[str]) -> Dict[str, Dict[str, Any]]:
    if not skus:
        return {}
    qs = ",".join(skus[:1000])
    try:
        return _get_json(f"{APP_BASE}/api/products/assets", {"skus": qs}) or {}
    except Exception:
        return {}

# ============================== Email helpers ================================
from email.message import EmailMessage
from email.utils import formatdate

def _validate_emails(s: str | List[str]) -> List[str]:
    if isinstance(s, list):
        emails = s
    else:
        emails = [x.strip() for x in (s or "").split(",") if x.strip()]
    if not emails:
        raise HTTPException(400, "No subscribers specified")
    bad = [e for e in emails if "@" not in e or "." not in e.split("@")[-1]]
    if bad:
        raise HTTPException(400, f"Invalid emails: {', '.join(bad)}")
    return emails

def _send_email(*, to_list: List[str], subject: str, body: str, attachments: List[Tuple[str, bytes]]) -> Dict[str, Any]:
    msg = EmailMessage()
    msg["From"] = SENDER
    msg["To"] = ", ".join(to_list)
    msg["Date"] = formatdate(localtime=True)
    msg["Subject"] = subject
    msg.set_content(body)

    for filename, data in attachments or []:
        msg.add_attachment(data, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)

    def save_to_outbox(reason: str) -> Dict[str, Any]:
        os.makedirs(OUTBOX_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(OUTBOX_DIR, f"out_{ts}.eml")
        with open(path, "wb") as f:
            f.write(msg.as_bytes())
        print(f"[email] saved to {os.path.abspath(path)} ({reason})")
        return {"transport": "file", "paths": [os.path.abspath(path)]}

    if not SMTP_HOST:
        return save_to_outbox("SMTP_HOST not configured")

    sec = SMTP_SECURITY
    if SMTP_PORT == 465 or (SMTP_SSL_FLAG and SMTP_SSL_FLAG.lower() in ("1","true","yes","on")):
        sec = "ssl"
    else:
        if SMTP_TLS_FLAG is not None:
            sec = "starttls" if SMTP_TLS_FLAG.lower() in ("1","true","yes","on") else "plain"

    try:
        if sec == "ssl":
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=SMTP_TIMEOUT, context=context) as s:
                if SMTP_DEBUG: s.set_debuglevel(1)
                s.ehlo()
                if SMTP_USER: s.login(SMTP_USER, SMTP_PASS)
                s.send_message(msg)
        elif sec == "starttls":
            context = ssl.create_default_context()
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=SMTP_TIMEOUT) as s:
                if SMTP_DEBUG: s.set_debuglevel(1)
                s.ehlo()
                s.starttls(context=context)
                s.ehlo()
                if SMTP_USER: s.login(SMTP_USER, SMTP_PASS)
                s.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=SMTP_TIMEOUT) as s:
                if SMTP_DEBUG: s.set_debuglevel(1)
                s.ehlo()
                if SMTP_USER: s.login(SMTP_USER, SMTP_PASS)
                s.send_message(msg)

        print(f"[email] sent via SMTP ({sec}) -> {to_list} host={SMTP_HOST}:{SMTP_PORT}")
        return {"transport": "smtp", "recipients": to_list, "security": sec}
    except Exception as e:
        print(f"[email] SMTP error: {e!r}; falling back to file.")
        return save_to_outbox(f"SMTP error: {e!r}")

# ====================== Comparison helpers + filtering + stats ===============
def _to_num(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s or s in ("n/a", "none", "null", "nan"):
        return None
    try:
        return float(s.replace(" ", "").replace(",", "."))
    except:
        return None

def _eff(promo, regular) -> Optional[float]:
    pv = _to_num(promo)
    return pv if pv is not None else _to_num(regular)

def _pivot_all(flat_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for r in flat_rows:
        sku = r.get("product_sku") or ""
        if not sku:
            continue
        g = out.setdefault(
            sku,
            {"name": r.get("product_name") or "N/A",
             "our": _eff(r.get("product_price_promo"), r.get("product_price_regular")),
             "comps": [], "sites_count": 0}
        )
        site = (r.get("competitor_site") or "").strip()
        price = _eff(r.get("competitor_price_promo"), r.get("competitor_price_regular"))
        url = r.get("competitor_url")
        cname = r.get("competitor_name")
        if site:
            g["comps"].append({"site": site, "price": price, "url": url, "name": cname})
    for g in out.values():
        g["sites_count"] = len({c["site"] for c in g["comps"]})
    return out

def _min_comp_info(comps: List[Dict[str, Any]]) -> Tuple[Optional[float], List[str], List[str]]:
    vals = [(c["price"], c["site"], c.get("url")) for c in comps if c["price"] is not None]
    if not vals:
        return (None, [], [])
    minv = min(v for v, _, _ in vals)
    sites = sorted({site for v, site, _ in vals if v == minv})
    urls = list({(url or "") for v, _, url in vals if v == minv and url})
    return (minv, sites, urls)

# ----------------------- PROMO filter: latest snapshot -----------------------
def _apply_only_promo_latest_snapshot(session, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Keep rows where the LATEST competitor snapshot for that site+competitor_sku has promo_price.
    Falls back to barcode if competitor_sku is missing.
    """
    if not rows:
        return rows

    # Map site code -> id
    site_codes = sorted({(r.get("competitor_site") or "").strip() for r in rows if (r.get("competitor_site") or "").strip()})
    code2id = {}
    if site_codes:
        q_sites = select(CompetitorSite.code, CompetitorSite.id).where(CompetitorSite.code.in_(site_codes))
        code2id = dict(get_session().__enter__().execute(q_sites).all())  # quick read-only context

    # Group competitor SKUs per site
    per_site_skus: Dict[int, List[str]] = {}
    for r in rows:
        code = (r.get("competitor_site") or "").strip()
        site_id = code2id.get(code)
        if not site_id:
            continue
        comp_sku = (r.get("competitor_sku") or "").strip()
        comp_bar = (r.get("competitor_barcode") or "").strip()
        key = comp_sku or comp_bar
        if key:
            per_site_skus.setdefault(site_id, []).append(key)

    if not per_site_skus:
        return []  # nothing we can verify => conservative: no promo rows

    # Build (site_id, sku) -> has_promo map using the *latest* snapshot per sku
    latest_has_promo: Dict[Tuple[int, str], bool] = {}
    for site_id, sku_list in per_site_skus.items():
        uniq = list({s for s in sku_list})
        # 1) latest ts per competitor_sku
        sub = (
            select(
                PriceSnapshot.competitor_sku.label("sku"),
                func.max(PriceSnapshot.ts).label("ts")
            )
            .where(
                PriceSnapshot.site_id == site_id,
                PriceSnapshot.competitor_sku.in_(uniq)
            )
            .group_by(PriceSnapshot.competitor_sku)
            .subquery()
        )
        # 2) fetch those latest rows to check promo_price
        q = (
            select(PriceSnapshot.competitor_sku, PriceSnapshot.promo_price)
            .join(sub,
                  (PriceSnapshot.competitor_sku == sub.c.sku) &
                  (PriceSnapshot.ts == sub.c.ts))
            .where(PriceSnapshot.site_id == site_id)
        )
        for sku, promo_price in get_session().__enter__().execute(q).all():
            latest_has_promo[(site_id, sku)] = (promo_price is not None)

    # Filter rows by that map
    out: List[Dict[str, Any]] = []
    for r in rows:
        code = (r.get("competitor_site") or "").strip()
        site_id = code2id.get(code)
        if not site_id:
            continue
        comp_sku = (r.get("competitor_sku") or "").strip()
        comp_bar = (r.get("competitor_barcode") or "").strip()
        key = comp_sku or comp_bar
        if not key:
            continue
        if latest_has_promo.get((site_id, key)):
            out.append(r)
    return out
# -----------------------------------------------------------------------------


def _apply_price_subset(rows: List[Dict[str, Any]], site_code: str, subset: PriceSubset) -> List[dict]:
    if subset == PriceSubset.all:
        return rows
    if site_code == "all":
        out = []
        grouped = _pivot_all(rows)
        for sku, g in grouped.items():
            our = g["our"]
            min_comp, _, _ = _min_comp_info(g["comps"])
            if our is None or min_comp is None:
                continue
            if subset == PriceSubset.ours_lower and our < min_comp:
                out.extend([r for r in rows if r.get("product_sku") == sku])
            elif subset == PriceSubset.ours_higher and our > min_comp:
                out.extend([r for r in rows if r.get("product_sku") == sku])
        return out
    out = []
    for r in rows:
        our = _eff(r.get("product_price_promo"), r.get("product_price_regular"))
        comp = _eff(r.get("competitor_price_promo"), r.get("competitor_price_regular"))
        if our is None or comp is None:
            continue
        if subset == PriceSubset.ours_lower and our < comp:
            out.append(r)
        elif subset == PriceSubset.ours_higher and our > comp:
            out.append(r)
    return out

# ---- price direction / category subtree / changed in last 24h ---------------
def _apply_price_direction(rows: List[Dict[str, Any]], direction: str) -> List[Dict[str, Any]]:
    direction = (direction or "any").lower()
    if direction == "any":
        return rows
    out: List[Dict[str, Any]] = []
    for r in rows:
        ours = _eff(r.get("product_price_promo"), r.get("product_price_regular"))
        comp = _eff(r.get("competitor_price_promo"), r.get("competitor_price_regular"))
        if ours is None or comp is None:
            continue
        if direction == "better" and ours < comp:
            out.append(r)
        elif direction == "worse" and ours > comp:
            out.append(r)
    return out

def _descendant_group_ids(session, root_id: int) -> List[int]:
    if not root_id:
        return []
    rows = session.execute(select(Group.id, Group.parent_id)).all()
    children: Dict[Optional[int], List[int]] = {}
    for gid, parent in rows:
        children.setdefault(parent, []).append(gid)
    out = [root_id]
    stack = [root_id]
    while stack:
        g = stack.pop()
        for c in children.get(g, []):
            out.append(c)
            stack.append(c)
    return out

def _apply_category_via_products(session, rows: List[Dict[str, Any]], category_id: int) -> List[Dict[str, Any]]:
    """
    Keep only rows whose product_sku belongs to a product with groupid in the selected subtree.
    Looks up Product.sku -> Product.groupid directly from DB (not the payload).
    """
    if not rows:
        return rows
    gids = set(_descendant_group_ids(session, int(category_id)))
    if not gids:
        return []
    skus = list({r.get("product_sku") for r in rows if r.get("product_sku")})
    if not skus:
        return []
    q = select(Product.sku, Product.groupid).where(Product.sku.in_(skus))
    sku2gid = dict(session.execute(q).all())
    keep = {sku for sku, gid in sku2gid.items() if gid in gids}
    return [r for r in rows if r.get("product_sku") in keep]

def _apply_changed_24h(session, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not rows:
        return rows
    cutoff = datetime.utcnow() - timedelta(hours=24)
    skus = list({r.get("product_sku") for r in rows if r.get("product_sku")})
    if not skus:
        return []
    q = select(Product.sku, Product.updated_at).where(Product.sku.in_(skus))
    found = dict(session.execute(q).all())
    recent = {sku for sku, upd in found.items() if upd and upd >= cutoff}
    return [r for r in rows if r.get("product_sku") in recent]
# -----------------------------------------------------------------------------


def _compose_stats(rows: List[Dict[str, Any]], site_code: str) -> Dict[str, int]:
    total = ours_lower = ours_higher = equal = no_comp = 0
    if site_code == "all":
        grouped = _pivot_all(rows)
        for _, g in grouped.items():
            total += 1
            our = g["our"]
            minc, _, _ = _min_comp_info(g["comps"])
            if our is None or minc is None:
                no_comp += 1
            elif our < minc:
                ours_lower += 1
            elif our > minc:
                ours_higher += 1
            else:
                equal += 1
    else:
        by_sku: Dict[str, Dict[str, Any]] = {}
        for r in rows:
            sku = r.get("product_sku") or ""
            if not sku:
                continue
            if sku not in by_sku:
                by_sku[sku] = r
        for r in by_sku.values():
            total += 1
            our = _eff(r.get("product_price_promo"), r.get("product_price_regular"))
            comp = _eff(r.get("competitor_price_promo"), r.get("competitor_price_regular"))
            if our is None or comp is None:
                no_comp += 1
            elif our < comp:
                ours_lower += 1
            elif our > comp:
                ours_higher += 1
            else:
                equal += 1
    return {"total": total, "ours_lower": ours_lower, "ours_higher": ours_higher, "equal": equal, "no_comp": no_comp}

def _setup_sheet(ws):
    ws.append([
        "SKU", "Нашата цена", "Сайт", "Конкурентна цена",
        "URL", "Сравнение", "Марка/Бележка", "Име"
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _head
        cell.alignment = _center
        cell.border = _border
    widths = [110, 110, 120, 120, 360, 120, 180, 460]
    for idx, px in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _px_to_w(px)

def _fill_sheet(ws, rows: List[Dict[str, Any]], site_code: str):
    for r in rows:
        sku = r.get("product_sku") or ""
        our_promo = r.get("product_price_promo")
        our_reg   = r.get("product_price_regular")
        our = _eff(our_promo, our_reg)
        comp_promo = r.get("competitor_price_promo")
        comp_reg   = r.get("competitor_price_regular")
        comp = _eff(comp_promo, comp_reg)
        site = r.get("competitor_site") or site_code or ""
        comp_url = r.get("competitor_url") or ""
        name = r.get("product_name") or ""
        label = r.get("competitor_label") or ""
        cmp_txt = "—"
        if our is not None and comp is not None:
            cmp_txt = "Нашата ↓" if our < comp else ("Нашата ↑" if our > comp else "Равни")
        ws.append([
            sku,
            our,
            site,
            comp,
            comp_url,
            cmp_txt,
            label,
            name
        ])
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for idx, cell in enumerate(r, start=1):
            cell.border = _border
            cell.alignment = _mid if idx in (1,2,8) else _right
        my = r[1].value
        cmpv = r[3].value
        if my is not None and cmpv is not None:
            if my < cmpv:
                r[5].fill = _green
            elif my > cmpv:
                r[5].fill = _red
    ws.freeze_panes = "A2"

def _build_report_workbook(rows: List[Dict[str, Any]], site_code: str, assets: Dict[str, Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    _setup_sheet(ws)
    _fill_sheet(ws, rows, site_code)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

def _union_rows_across_tags(tag_rows: List[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    seen = set()
    out: List[Dict[str, Any]] = []
    for rows in tag_rows:
        for r in rows:
            key = (
                r.get("product_sku") or "",
                r.get("competitor_site") or "",
                r.get("competitor_sku") or "",
                r.get("competitor_barcode") or "",
            )
            if key in seen:
                continue
            seen.add(key)
            out.append(r)
    return out

# ============================ Build + send logic ==============================
def _build_and_send(rule: EmailRule) -> Dict[str, Any]:
    emails = _validate_emails(rule.subscribers)

    total_all = ours_lower_all = ours_higher_all = equal_all = no_comp_all = 0
    tag_ids_raw = rule.tag_ids or [None]

    per_tag_rows: List[List[Dict[str, Any]]] = []
    for tag_id in tag_ids_raw:
        rows = _fetch_compare(rule.site_code or "all", tag_id, rule.brand, q=None)

        # --- promo only: latest snapshot check
        if rule.only_promo:
            with get_session() as _ses_p:
                rows = _apply_only_promo_latest_snapshot(_ses_p, rows)

        rows = _apply_price_subset(rows, rule.site_code or "all", rule.price_subset)

        # --- category via products.groupid subtree
        if getattr(rule, "category_id", None):
            with get_session() as _ses_cat:
                rows = _apply_category_via_products(_ses_cat, rows, int(rule.category_id))

        # --- price direction any|better|worse
        rows = _apply_price_direction(rows, getattr(rule, "price_direction", "any"))

        # --- changed in last 24h (our product updated_at)
        if getattr(rule, "changed_24h", False):
            with get_session() as _ses_ch:
                rows = _apply_changed_24h(_ses_ch, rows)

        stats = _compose_stats(rows, rule.site_code or "all")
        total_all       += stats["total"]
        ours_lower_all  += stats["ours_lower"]
        ours_higher_all += stats["ours_higher"]
        equal_all       += stats["equal"]
        no_comp_all     += stats["no_comp"]

        per_tag_rows.append(rows)

    if len(per_tag_rows) > 1:
        final_rows = _union_rows_across_tags(per_tag_rows)
        label_for_name = "combined"
    else:
        final_rows = per_tag_rows[0] if per_tag_rows else []
        label_for_name = "all" if tag_ids_raw == [None] else str(tag_ids_raw[0])

    skus = [r.get("product_sku") for r in final_rows if r.get("product_sku")]
    assets = _fetch_assets(list(dict.fromkeys(skus))) if skus else {}
    data = _build_report_workbook(final_rows, rule.site_code or "all", assets)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    attachments = [(f"report_{label_for_name}_{ts}.xlsx", data)]

    body_lines = [
        f"Report: {rule.name}",
        f"Site scope: {rule.site_code or 'all'}",
        f"Price subset: {rule.price_subset.value}",
        f"Only promo: {'yes' if rule.only_promo else 'no'}",
        f"Brand filter: {rule.brand or '—'}",
        f"Tags: {', '.join([str(t) for t in (rule.tag_ids or ['all'])])}",
        f"Category: {getattr(rule, 'category_id', None) or '—'}",
        f"Price direction: {getattr(rule, 'price_direction', 'any')}",
        f"Changed last 24h: {'yes' if getattr(rule, 'changed_24h', False) else 'no'}",
        "",
        "Summary:",
        f"  • Items: {total_all}",
        f"  • Ours lower: {ours_lower_all}",
        f"  • Ours higher: {ours_higher_all}",
        f"  • Equal: {equal_all}",
        f"  • Missing competitor price: {no_comp_all}",
        "",
        "This is an automated message.",
    ]
    info = _send_email(
        to_list=_validate_emails(rule.subscribers),
        subject=f"[Price Report] {rule.name}",
        body="\n".join(body_lines),
        attachments=attachments,
    )
    return info

# ============================= CRUD & scheduling ==============================

def _enum_value(v, default="any") -> str:
    if v is None:
        return default
    val = getattr(v, "value", None)
    if isinstance(val, str):
        return val
    s = str(v)
    if "." in s:
        s = s.split(".")[-1]
    return s

@router.get("/email/rules", response_model=List[EmailRuleOut])
def list_rules():
    with get_session() as session:
        rows = list(session.execute(select(EmailRule).order_by(EmailRule.created_on.desc())).scalars())
        return rows

@router.post("/email/rules", response_model=EmailRuleOut)
def create_rule(payload: EmailRuleIn):
    _validate_emails(payload.subscribers)
    with get_session() as session:
        row = EmailRule(
            name=payload.name,
            tag_ids=payload.tag_ids or [],
            brand=payload.brand,
            site_code=payload.site_code or "all",
            price_subset=PriceSubset(payload.price_subset.value),
            only_promo=bool(getattr(payload, "only_promo", False) or getattr(payload, "promo_only", False)),
            category_id=getattr(payload, "category_id", None),
            price_direction=_enum_value(getattr(payload, "price_direction", "any")),
            changed_24h=bool(getattr(payload, "changed_24h", False)),
            subscribers=payload.subscribers,
            notes=payload.notes,
            created_by="UI",
        )
        session.add(row)
        session.commit()
        session.refresh(row)
        return row

@router.put("/email/rules/{rule_id}", response_model=EmailRuleOut)
def update_rule(rule_id: int, payload: EmailRuleIn):
    _validate_emails(payload.subscribers)
    with get_session() as session:
        row = session.get(EmailRule, rule_id)
        if not row:
            raise HTTPException(404, "Rule not found")
        row.name = payload.name
        row.tag_ids = payload.tag_ids or []
        row.brand = payload.brand
        row.site_code = payload.site_code or "all"
        row.price_subset = PriceSubset(payload.price_subset.value)
        row.only_promo = bool(getattr(payload, "only_promo", False) or getattr(payload, "promo_only", False))
        row.category_id = getattr(payload, "category_id", None)
        row.price_direction = _enum_value(getattr(payload, "price_direction", "any"))
        row.changed_24h = bool(getattr(payload, "changed_24h", False))
        row.subscribers = payload.subscribers
        row.notes = payload.notes
        session.commit()
        session.refresh(row)
        return row

@router.delete("/email/rules/{rule_id}")
def delete_rule(rule_id: int):
    with get_session() as session:
        row = session.get(EmailRule, rule_id)
        if not row:
            raise HTTPException(404, "Rule not found")
        session.delete(row)
        session.commit()
        return {"ok": True}

@router.get("/email/schedule", response_model=WeeklySchedule)
def get_schedule():
    with get_session() as session:
        row = session.execute(select(EmailWeeklySchedule)).scalars().first()
        if not row:
            row = EmailWeeklySchedule()
            session.add(row)
            session.commit()
            session.refresh(row)
        data = row.data or {}
        return WeeklySchedule(**{
            "mon": data.get("mon"), "tue": data.get("tue"), "wed": data.get("wed"),
            "thu": data.get("thu"), "fri": data.get("fri"), "sat": data.get("sat"), "sun": data.get("sun"),
        })

@router.put("/email/schedule", response_model=WeeklySchedule)
def put_schedule(payload: WeeklySchedule):
    with get_session() as session:
        row = session.execute(select(EmailWeeklySchedule)).scalars().first()
        if not row:
            row = EmailWeeklySchedule()
            session.add(row)
        row.data = payload.dict()
        row.next_run_at = None
        session.commit()
        session.refresh(row)
        return payload

@router.post("/email/schedule", response_model=WeeklySchedule)
def post_schedule(payload: WeeklySchedule):
    return put_schedule(payload)

@router.post("/email/send/{rule_id}")
def send_now(rule_id: int):
    with get_session() as session:
        row = session.get(EmailRule, rule_id)
        if not row:
            raise HTTPException(404, "Rule not found")
    info = _build_and_send(row)
    return {"ok": info.get("transport") == "smtp", **info}

@router.post("/email/send")
def send_now_query(rule_id: int = Query(...)):
    return send_now(rule_id)

@router.post("/email/send-all")
def send_all_now():
    with get_session() as session:
        rules = list(session.execute(select(EmailRule)).scalars())
    results = []
    for r in rules:
        try:
            info = _build_and_send(r)
            results.append({"rule_id": r.id, "name": r.name, **info})
        except Exception as e:
            results.append({"rule_id": r.id, "name": r.name, "error": str(e)})
    return {"ok": True, "results": results}

# ============================== Simple scheduler =============================
async def _compute_next_run(now: datetime, schedule: dict) -> Optional[datetime]:
    mapday = ["mon","tue","wed","thu","fri","sat","sun"]
    for delta in range(0, 8):
        dt = now + timedelta(days=delta)
        key = mapday[dt.weekday()]
        hhmm = (schedule or {}).get(key)
        if not hhmm or hhmm in ("-","--",""):
            continue
        try:
            hh, mm = [int(x) for x in hhmm.split(":")]
        except:
            continue
        cand = dt.replace(hour=hh, minute=mm, second=0, microsecond=0)
        if cand >= now:
            return cand
    return None

async def email_scheduler_loop():
    while True:
        try:
            with get_session() as session:
                row = session.execute(select(EmailWeeklySchedule)).scalars().first()
                if not row:
                    row = EmailWeeklySchedule()
                    session.add(row)
                    session.commit()
                    session.refresh(row)
                now = datetime.now()
                if not row.next_run_at:
                    row.next_run_at = await _compute_next_run(now, row.data or {})
                    session.commit()
                if row.next_run_at and now >= row.next_run_at:
                    rules = list(session.execute(select(EmailRule)).scalars())
                    for r in rules:
                        try:
                            _build_and_send(r)
                        except Exception as e:
                            print("scheduler rule error:", r.id, e)
                    row.next_run_at = await _compute_next_run(now + timedelta(seconds=1), row.data or {})
                    session.commit()
        except Exception as e:
            print("email scheduler loop exception:", e)
        await asyncio.sleep(60)
