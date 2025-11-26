# app/services/erp_import.py
# -*- coding: utf-8 -*-
from __future__ import annotations

from typing import BinaryIO, Dict, List, Optional
import io
import os
from datetime import datetime
import textwrap

import requests
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from sqlalchemy import select

from app.db import get_session
from app.models import Product

# ---------------- Zeron config ----------------

ZERON_URL = os.getenv(
    "ZERON_URL",
    "https://sysserver.praktis.bg:37005/ZeronServerService/DataExchange",
)
ZERON_DATABASE = os.getenv("ZERON_DB", "zdbMegadom2023")
ZERON_USERCODE = os.getenv("ZERON_USERCODE", "1831")
ZERON_USERPASS = os.getenv("ZERON_USERPASS", "angelbangel33")
ZERON_STOREHOUSE = os.getenv("ZERON_STOREHOUSE", "1001")

# how many SKUs per Zeron request
ZERON_MAX_PER_REQUEST = int(os.getenv("ZERON_MAX_PER_REQUEST", "200"))
ZERON_TIMEOUT = int(os.getenv("ZERON_TIMEOUT", "60"))
ZERON_VERIFY_SSL = os.getenv("ZERON_SSL_VERIFY", "1").lower() in ("1", "true", "yes", "on")

# ---------------- Excel helpers ----------------

def _normalize_header(value: str) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    while "  " in s:
        s = s.replace("  ", " ")
    return s

# headers we treat as the SKU column (case-insensitive, space-insensitive)
SKU_HEADER_KEYWORDS = (
    # Bulgarian variants
    "код",          # "Код"
    "ску",          # "Ску"
    "код на zeron",
    "код в zeron",
    "код zeron",
    # English variants
    "sku",
    "skus",
    "codes",
)

def _is_sku_header_cell(text: str) -> bool:
    if not text:
        return False
    norm = _normalize_header(text)

    # exact simple names
    if norm in ("код", "ску", "sku", "skus", "codes"):
        return True

    # keyword based
    for kw in SKU_HEADER_KEYWORDS:
        if kw in norm:
            return True
    return False

def _normalize_sku_cell(value) -> Optional[str]:
    if value is None:
        return None
    # Numbers (e.g. 35566672 or 35566672.0)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    if not s:
        return None
    # Excelish "35566672.0"
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s

def extract_skus_from_excel(f: BinaryIO) -> List[str]:
    """
    Reads an Excel file and returns a list of SKUs based on header names.

    Recognised header names / patterns (case-insensitive):
      - Код, Ску
      - Код на Zeron, Код в Zeron, Код Zeron
      - Sku, skus, codes
    """
    f.seek(0)
    wb = load_workbook(f, read_only=True, data_only=True)
    skus: List[str] = []
    seen: set[str] = set()

    for sheet in wb.worksheets:
        sku_col_idx: Optional[int] = None
        header_row_idx: Optional[int] = None

        # Search header row (first 20 rows max)
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
            for cell in row:
                text = str(cell.value).strip() if cell.value is not None else ""
                if _is_sku_header_cell(text):
                    sku_col_idx = cell.column  # 1-based
                    header_row_idx = cell.row
                    break
            if sku_col_idx is not None:
                break

        if sku_col_idx is None or header_row_idx is None:
            continue  # sheet without recognised header

        # Data rows
        for row in sheet.iter_rows(min_row=header_row_idx + 1, max_row=sheet.max_row):
            cell = row[sku_col_idx - 1]  # openpyxl columns are 1-based
            sku = _normalize_sku_cell(cell.value)
            if not sku:
                continue
            if sku in seen:
                continue
            seen.add(sku)
            skus.append(sku)

    if not skus:
        raise ValueError(
            "Не успях да намеря колона със SKU. Очаквам заглавие като "
            "\"Код\", \"Код в Zeron\", \"Sku\", \"Skus\", \"Codes\" и т.н."
        )
    return skus

# ---------------- Zeron XML helpers ----------------

def _build_zeron_payload(skus: List[str]) -> str:
    rows = [f"<Row><InvCode>{sku}</InvCode></Row>" for sku in skus]
    invlist_xml = "".join(rows)
    xml = f"""
<Data>
  <Destination>
    <Database>{ZERON_DATABASE}</Database>
    <UserCode>{ZERON_USERCODE}</UserCode>
    <UserPass>{ZERON_USERPASS}</UserPass>
    <Operation>
      <OperName>ESLGetInvPriceGroupsPrices</OperName>
      <Parameters>
        <InvList>
          {invlist_xml}
        </InvList>
        <StoreHouse>{ZERON_STOREHOUSE}</StoreHouse>
      </Parameters>
    </Operation>
  </Destination>
</Data>
""".strip()
    # compact for smaller payload
    return textwrap.dedent(xml).replace("\n", "").replace("  ", "")

def _parse_zeron_response(xml_text: str) -> Dict[str, dict]:
    """
    Parse Zeron XML into a dict keyed by InvCode (sku).
    """
    result: Dict[str, dict] = {}
    root = ET.fromstring(xml_text)

    success_text = (root.findtext(".//Success") or "").strip().lower()
    if success_text not in ("true", "1", "yes"):
        msg = (root.findtext(".//ErrorMessage") or "").strip()
        raise RuntimeError(f"Zeron returned error: {msg or 'Unknown error'}")

    for table_el in root.findall(".//Table"):
        def get(tag: str) -> Optional[str]:
            el = table_el.find(tag)
            return el.text.strip() if el is not None and el.text is not None else None

        inv_code = get("InvCode")
        if not inv_code:
            continue

        price_raw = get("Price")
        try:
            price = float(price_raw.replace(",", ".")) if price_raw else None
        except Exception:
            price = None

        group_raw = get("GroupID")
        try:
            group_id = int(group_raw) if group_raw else None
        except Exception:
            group_id = None

        result[inv_code] = {
            "sku": inv_code,
            "barcode": get("Barcode"),
            "name": get("InvName"),
            "measure": get("Measure"),
            "price_group_type": get("PriceGroupType"),
            "price_group_code": get("PriceGroupCode"),
            "price": price,
            "currency_code": get("CurrencyCode"),
            "from_date": get("FromDate"),
            "allow_better_prices": get("AllowBetterPrices"),
            "in_brochure": get("InBroschure"),
            "on_stock": get("OnStock"),
            "blocked_delivery": get("BlockedDelivery"),
            "groupid": group_id,
            "brand": get("Trademark"),
            "item_number": get("SuppItemCode"),
        }
    return result

def fetch_zeron_for_skus(all_skus: List[str]) -> Dict[str, dict]:
    """
    Call Zeron in chunks and return merged result mapping sku -> info.
    """
    all_skus = [s for s in all_skus if s]
    merged: Dict[str, dict] = {}
    if not all_skus:
        return merged

    for i in range(0, len(all_skus), ZERON_MAX_PER_REQUEST):
        chunk = all_skus[i:i + ZERON_MAX_PER_REQUEST]
        payload = _build_zeron_payload(chunk)
        r = requests.post(
            ZERON_URL,
            data=payload.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=ZERON_TIMEOUT,
            verify=ZERON_VERIFY_SSL,
        )
        r.raise_for_status()
        part = _parse_zeron_response(r.text)
        merged.update(part)
    return merged

# ---------------- DB upsert helpers ----------------

def upsert_products_from_erp(data: Dict[str, dict]) -> Dict[str, int]:
    """
    Insert / update Product rows from ERP data.
    Returns stats: created / updated.
    """
    if not data:
        return {"created": 0, "updated": 0}

    skus = list(data.keys())
    created = 0
    updated = 0

    with get_session() as session:
        existing = {
            p.sku: p
            for p in session.execute(
                select(Product).where(Product.sku.in_(skus))
            ).scalars()
        }

        for sku, info in data.items():
            p = existing.get(sku)
            if p is None:
                p = Product(sku=sku)
                session.add(p)
                created += 1
            else:
                updated += 1

            # Map ERP fields to Product columns
            if info.get("barcode") is not None:
                p.barcode = info["barcode"] or p.barcode
            if info.get("item_number") is not None:
                p.item_number = info["item_number"] or p.item_number
            if info.get("brand") is not None:
                p.brand = info["brand"] or p.brand
            if info.get("name") is not None:
                p.name = info["name"] or p.name

            # price_regular from ERP
            if "price" in info:
                p.price_regular = info["price"]

            # groupid is FK to groups; assumes IDs are already imported from ERP
            groupid = info.get("groupid")
            if groupid:
                p.groupid = groupid

            p.updated_at = datetime.utcnow()

        session.commit()

    return {"created": created, "updated": updated}

# ---------------- Orchestrators ----------------

def import_excel_and_update_products(f: BinaryIO) -> Dict[str, object]:
    """
    High-level helper used by the /erp/import_excel endpoint.
    """
    skus = extract_skus_from_excel(f)
    erp_data = fetch_zeron_for_skus(skus)
    stats = upsert_products_from_erp(erp_data)
    missing = sorted(set(skus) - set(erp_data.keys()))
    return {
        "ok": True,
        "skus_in_file": len(skus),
        "skus_unique": len(set(skus)),
        "skus_with_data": len(erp_data),
        "created": stats["created"],
        "updated": stats["updated"],
        "missing_in_erp": missing,
    }

def refresh_all_products_once() -> Dict[str, object]:
    """
    Used by daily cronjob: refresh all existing products from ERP.
    """
    with get_session() as session:
        rows = session.execute(select(Product.sku)).all()
        skus = [r[0] for r in rows if r[0]]

    skus = sorted(set(skus))
    if not skus:
        return {"ok": True, "total_skus": 0, "skus_with_data": 0, "created": 0, "updated": 0}

    erp_data = fetch_zeron_for_skus(skus)
    stats = upsert_products_from_erp(erp_data)
    return {
        "ok": True,
        "total_skus": len(skus),
        "skus_with_data": len(erp_data),
        "created": stats["created"],
        "updated": stats["updated"],
    }
