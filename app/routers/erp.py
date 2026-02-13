# app/routers/erp.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import os
import logging
from datetime import datetime
from typing import Dict, List, Optional

import requests
import xmltodict
from fastapi import APIRouter, UploadFile, File, HTTPException
from sqlalchemy import select
from openpyxl import load_workbook

from app.db import get_session
from app.models import Product, Group

# NOTE:
# main.py mounts this router with:
#   app.include_router(r_erp.router, prefix="/api", tags=["erp"])
#
# So a route defined as "/erp/import_excel" here becomes:
#   /api/erp/import_excel

router = APIRouter()
logger = logging.getLogger(__name__)

# =============================================================================
# Zeron connection config
# =============================================================================

ZERON_URL = os.getenv(
    "ZERON_URL",
    "https://sysserver.praktis.bg:37005/ZeronServerService/DataExchange",
)
ZERON_DATABASE = os.getenv("ZERON_DB", "zdbMegadom")
ZERON_USERCODE = os.getenv("ZERON_USERCODE", "1831")
ZERON_USERPASS = os.getenv("ZERON_USERPASS", "angelbangel33")
ZERON_STOREHOUSE = os.getenv("ZERON_STOREHOUSE", "1001")

ZERON_MAX_PER_REQUEST = int(os.getenv("ZERON_MAX_PER_REQUEST", "200"))
ZERON_TIMEOUT = int(os.getenv("ZERON_TIMEOUT", "60"))
ZERON_VERIFY_SSL = os.getenv("ZERON_SSL_VERIFY", "1").lower() in ("1", "true", "yes", "on")


# =============================================================================
# Helpers: Excel parsing (find SKU column by header name)
# =============================================================================

def _normalize_header(value: str) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    while "  " in s:
        s = s.replace("  ", " ")
    return s


# Names / fragments for the SKU column header
SKU_HEADER_KEYWORDS = (
    # Bulgarian variants
    "код",              # "Код"
    "ску",              # "Ску"
    "код на zeron",
    "код в zeron",
    "код zeron",
    "ков зерон",        # typo: "Ков Зерон"
    # English variants
    "sku",
    "skus",
    "codes",
)


def _is_sku_header_cell(text: str) -> bool:
    if not text:
        return False
    norm = _normalize_header(text)

    # exact simple names
    if norm in ("код", "ску", "sku", "skus", "codes"):
        return True

    for kw in SKU_HEADER_KEYWORDS:
        if kw in norm:
            return True
    return False


def _normalize_sku_cell(value) -> Optional[str]:
    if value is None:
        return None
    # Numeric cells (e.g. 35566672 or 35566672.0)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    if not s:
        return None
    # Excel-ish "35566672.0"
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def extract_skus_from_excel(data: bytes) -> List[str]:
    """
    Reads an Excel file and returns a list of SKUs based on header names.

    Recognised header names / patterns (case-insensitive):
      - Код, Ску
      - Код на Zeron, Код в Zeron, Код Zeron, Ков Зерон
      - Sku, skus, codes
    """
    f = io.BytesIO(data)
    wb = load_workbook(f, read_only=True, data_only=True)
    skus: List[str] = []
    seen: set[str] = set()

    for sheet in wb.worksheets:
        sku_col_idx: Optional[int] = None
        header_row_idx: Optional[int] = None

        # Search header row (first 20 rows max)
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
            for cell in row:
                text = str(cell.value).strip() if cell.value is not None else ""
                if _is_sku_header_cell(text):
                    sku_col_idx = cell.column  # 1-based index
                    header_row_idx = cell.row
                    break
            if sku_col_idx is not None:
                break

        if sku_col_idx is None or header_row_idx is None:
            continue  # sheet without recognised header

        # Data rows
        for row in sheet.iter_rows(min_row=header_row_idx + 1, max_row=sheet.max_row):
            cell = row[sku_col_idx - 1]  # openpyxl columns are 1-based
            sku = _normalize_sku_cell(cell.value)
            if not sku:
                continue
            if sku in seen:
                continue
            seen.add(sku)
            skus.append(sku)

    if not skus:
        raise ValueError(
            "Не успях да намеря колона със SKU. Очаквам заглавие като "
            '"Код", "Код в Zeron", "Код на Зерон", "Sku", "Skus", "Codes" и т.н.'
        )
    return skus


# =============================================================================
# Common helper (taken from your code – ONLY used for price logic)
# =============================================================================

class Common:
    @staticmethod
    def process_string(s):
        return (str(s).strip() if s is not None else "")

    @staticmethod
    def calculate_discounted_price(price: float, discount: float) -> float:
        return price * (100.0 - discount) / 100.0

    @classmethod
    def process_entry(cls, entry, all_entries):
        if 'Discount' in entry:
            entry['DiscountedPrice'] = cls.calculate_discounted_price(float(entry['Price']), float(entry['Discount']))
        else:
            # Find the maximum price for the same InvCode to determine the base price
            inv_code = entry['InvCode']
            base_price = max(float(e['Price']) for e in all_entries if e['InvCode'] == inv_code)

            discounted_price = float(entry['Price'])  # The current price is the discounted price

            if base_price > discounted_price:
                discount_percentage = round(((base_price - discounted_price) / base_price) * 100)
                entry['Discount'] = discount_percentage  # Store discount percentage

            entry['DiscountedPrice'] = discounted_price  # Store the actual discounted price
            entry['Price'] = str(base_price)  # Ensure that the base price is stored correctly


# =============================================================================
# Helpers: Zeron XML generation + parsing + YOUR selection logic
# =============================================================================

def _build_zeron_payload(skus: List[str]) -> str:
    rows = "".join(f"<Row><InvCode>{sku}</InvCode></Row>" for sku in skus)
    xml = f"""
<Data>
  <Destination>
    <Database>{ZERON_DATABASE}</Database>
    <UserCode>{ZERON_USERCODE}</UserCode>
    <UserPass>{ZERON_USERPASS}</UserPass>
    <Operation>
      <OperName>GetPriceCheckerData</OperName>
      <Parameters>
        <InvList>
          {rows}
        </InvList>
        <StoreHouse>{ZERON_STOREHOUSE}</StoreHouse>
      </Parameters>
    </Operation>
  </Destination>
</Data>
""".strip()
    # compact payload
    return xml.replace("\n", "").replace("  ", "")


def _parse_zeron_response(xml_text: str) -> Dict[str, dict]:
    """
    Parse Zeron XML into a dict keyed by InvCode (sku), but the selection
    of which row to use per SKU uses EXACTLY your if/else logic.
    """
    # Your code uses xmltodict, so we'll do the same
    new_data = xmltodict.parse(xml_text)

    selected_tables = []
    inv_codes: Dict[str, List[dict]] = {}

    # ---- YOUR LOGIC START (copied as-is, only "url / headers / data" part omitted) ----
    # Check if 'Table' key exists in response
    if new_data['Response']['Destination']['Operation']['DataSet'] is not None:
        logger.info("Successfully received Zeron data.")

        tables = new_data['Response']['Destination']['Operation']['DataSet']['Table']
        if not isinstance(tables, list):
            tables = [tables]

        # Group entries by InvCode
        for table in tables:
            entry = dict(table)
            entry["Storehouse"] = ZERON_STOREHOUSE
            inv_code = entry.get('InvCode')
            entry['CurrencyCode'] = Common.process_string(entry['CurrencyCode'])

            if inv_code:
                # Separating the repeating inv_codes
                if inv_code not in inv_codes:
                    inv_codes[inv_code] = []
                inv_codes[inv_code].append(entry)
    else:
        logger.error("Zeron data response has empty DataSet.")

    # Process each group of entries with the same InvCode
    for inv_code, entries in inv_codes.items():
        allow_better_prices_0_entries = [entry for entry in entries if entry['AllowBetterPrices'] == '0']

        if allow_better_prices_0_entries:
            for entry in allow_better_prices_0_entries:
                Common.process_entry(entry, entries)
            # If we have entries with AllowBetterPrices 0, select the one with the lowest PriceGroupCode
            allow_better_prices_0_entries.sort(key=lambda x: int(x['PriceGroupCode']))
            selected_tables.append(allow_better_prices_0_entries[0])
        else:
            allow_better_prices_2_entries = [entry for entry in entries if entry['AllowBetterPrices'] == '2']

            if allow_better_prices_2_entries:
                # If we have entries with AllowBetterPrices 2, filter by PriceGroupType 10 and 20, choose the lowest price
                allow_better_prices_2_filtered = [
                    entry for entry in entries if entry['PriceGroupType'] in ['10', '20']
                ]
                if allow_better_prices_2_filtered:
                    for entry in allow_better_prices_2_filtered:
                        Common.process_entry(entry, entries)
                    allow_better_prices_2_filtered.sort(key=lambda x: float(x['DiscountedPrice']))
                    selected_tables.append(allow_better_prices_2_filtered[0])
            else:
                allow_better_prices_3_entries = [entry for entry in entries if entry['AllowBetterPrices'] == '3']

                if allow_better_prices_3_entries:
                    # If we have entries with AllowBetterPrices 3, filter by PriceGroupType 20 and 30, choose the lowest price
                    allow_better_prices_3_filtered = [
                        entry for entry in entries if entry['PriceGroupType'] in ['20', '30']
                    ]
                    if allow_better_prices_3_filtered:
                        for entry in allow_better_prices_3_filtered:
                            Common.process_entry(entry, entries)
                        allow_better_prices_3_filtered.sort(key=lambda x: float(x['DiscountedPrice']))
                        selected_tables.append(allow_better_prices_3_filtered[0])
                else:
                    allow_better_prices_1_entries = [entry for entry in entries if entry['AllowBetterPrices'] == '1']

                if allow_better_prices_1_entries:
                    # Only calculate discounts for PriceGroupType != '10'
                    for entry in allow_better_prices_1_entries:
                        if entry['PriceGroupType'] != '10':
                            Common.process_entry(entry, entries)
                        else:
                            # keep the price as-is and zero out any discount
                            entry['DiscountedPrice'] = float(entry['Price'])
                            entry['Discount'] = 0

                    # now pick the one with the lowest discounted price
                    allow_better_prices_1_entries.sort(
                        key=lambda x: float(x['DiscountedPrice'])
                    )
                    selected_tables.append(allow_better_prices_1_entries[0])
    # ---- YOUR LOGIC END ----

    # Now we have selected_tables: one chosen row per InvCode.
    # Convert to the dict format used by upsert_products_from_erp()
    result: Dict[str, dict] = {}

    for entry in selected_tables:
        inv_code = entry.get("InvCode")
        if not inv_code:
            continue

        price_raw = entry.get("Price")
        try:
            price = float(str(price_raw).replace(",", ".")) if price_raw else None
        except Exception:
            price = None

        group_raw = entry.get("GroupID")
        try:
            groupid = int(group_raw) if group_raw else None
        except Exception:
            groupid = None

        result[inv_code] = {
            "sku": inv_code,
            "barcode": entry.get("Barcode"),
            "name": entry.get("InvName"),
            "measure": entry.get("Measure"),
            "price_group_type": entry.get("PriceGroupType"),
            "price_group_code": entry.get("PriceGroupCode"),
            "price": price,
            "currency_code": entry.get("CurrencyCode"),
            "from_date": entry.get("FromDate"),
            "allow_better_prices": entry.get("AllowBetterPrices"),
            "in_brochure": entry.get("InBroschure"),
            "on_stock": entry.get("OnStock"),
            "blocked_delivery": entry.get("BlockedDelivery"),
            "groupid": groupid,
            "brand": entry.get("Trademark"),
            "item_number": entry.get("SuppItemCode"),
        }

    return result


def fetch_zeron_for_skus(all_skus: List[str]) -> Dict[str, dict]:
    """
    Call Zeron in chunks and return merged result mapping sku -> info,
    after applying your selection logic.
    """
    skus = [s for s in all_skus if s]
    merged: Dict[str, dict] = {}
    if not skus:
        return merged

    for i in range(0, len(skus), ZERON_MAX_PER_REQUEST):
        chunk = skus[i:i + ZERON_MAX_PER_REQUEST]
        payload = _build_zeron_payload(chunk)
        r = requests.post(
            ZERON_URL,
            data=payload.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=ZERON_TIMEOUT,
            verify=ZERON_VERIFY_SSL,
        )
        r.raise_for_status()
        part = _parse_zeron_response(r.text)
        merged.update(part)
    return merged


# =============================================================================
# Helpers: DB upsert
# =============================================================================

def upsert_products_from_erp(data: Dict[str, dict]) -> Dict[str, int]:
    """
    Insert / update Product rows from ERP data.
    Returns stats: created / updated.
    """
    if not data:
        return {"created": 0, "updated": 0}

    skus = list(data.keys())
    created = 0
    updated = 0

    with get_session() as session:
        # existing products
        existing = {
            p.sku: p
            for p in session.execute(
                select(Product).where(Product.sku.in_(skus))
            ).scalars()
        }

        # **new**: load valid group ids so we don't violate the FK
        valid_group_ids = {
            gid for (gid,) in session.execute(select(Group.id)).all()
        }

        for sku, info in data.items():
            p = existing.get(sku)
            if p is None:
                p = Product(sku=sku)
                session.add(p)
                existing[sku] = p
                created += 1
            else:
                updated += 1

            # Map ERP fields to Product columns
            if info.get("barcode") is not None:
                p.barcode = info["barcode"] or p.barcode
            if info.get("item_number") is not None:
                p.item_number = info["item_number"] or p.item_number
            if info.get("brand") is not None:
                p.brand = info["brand"] or p.brand
            if info.get("name") is not None:
                p.name = info["name"] or p.name

            # price_regular from ERP (selected row)
            if "price" in info:
                p.price_regular = info["price"]

            # groupid is FK to groups; **only set if it exists in groups.id**
            groupid = info.get("groupid")
            if groupid and groupid in valid_group_ids:
                p.groupid = groupid
            # else: leave p.groupid as-is to avoid FK violation

            p.updated_at = datetime.utcnow()

        session.commit()

    return {"created": created, "updated": updated}


# =============================================================================
# Endpoints
# =============================================================================

@router.post("/erp/ingest_xml")
async def ingest_xml(xml_file: UploadFile = File(...)):
    """
    OLD behaviour: upload XML (already exported from Zeron) and import products.
    Kept so you don't lose previous functionality.

    Frontend (current erp.html) sends field name: "xml_file"
    to /api/erp/ingest_xml.
    """
    try:
        raw = await xml_file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не мога да прочета файла: {e}")

    # try UTF-8, fallback to Windows-1251
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("cp1251", errors="ignore")

    try:
        data = _parse_zeron_response(text)
        stats = upsert_products_from_erp(data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Грешен XML или неочакван формат: {e}")

    return {
        "ok": True,
        "rows_in_xml": len(data),
        "created": stats["created"],
        "updated": stats["updated"],
    }


@router.post("/erp/import_excel")
async def import_erp_excel(file: UploadFile = File(...)):
    """
    NEW behaviour: upload Excel with SKUs; call Zeron and upsert into products.

    Excel може да има допълнителни колони – търсим колоната с име:
      „Код“, „Код в Zeron“, „Код на Зерон“, „Sku“, „Skus“, „Codes“, „Ков Зерон“ и т.н.

    Final path (with main.py prefix) is:
      POST /api/erp/import_excel
    """
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Моля, качете Excel файл (.xlsx или .xls).")

    content = await file.read()

    try:
        skus = extract_skus_from_excel(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Грешка при четене на Excel: {e}")

    try:
        erp_data = fetch_zeron_for_skus(skus)
        stats = upsert_products_from_erp(erp_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при заявка към Zeron: {e}")

    missing = sorted(set(skus) - set(erp_data.keys()))

    return {
        "ok": True,
        "skus_in_file": len(skus),
        "skus_unique": len(set(skus)),
        "skus_with_data": len(erp_data),
        "created": stats["created"],
        "updated": stats["updated"],
        "missing_in_erp": missing,
    }


@router.post("/erp/refresh_all")
async def erp_refresh_all_products():
    """
    Endpoint intended to be called by external cronjob (no async loop inside app).

    Fetches ALL product.sku from DB, calls Zeron, and updates them.

    Final path (with main.py prefix) is:
      POST /api/erp/refresh_all
    """
    with get_session() as session:
        rows = session.execute(select(Product.sku)).all()
        all_skus = [r[0] for r in rows if r[0]]

    all_skus = sorted(set(all_skus))
    if not all_skus:
        return {
            "ok": True,
            "total_skus": 0,
            "skus_with_data": 0,
            "created": 0,
            "updated": 0,
        }

    try:
        erp_data = fetch_zeron_for_skus(all_skus)
        stats = upsert_products_from_erp(erp_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Грешка при заявка към Zeron: {e}")

    return {
        "ok": True,
        "total_skus": len(all_skus),
        "skus_with_data": len(erp_data),
        "created": stats["created"],
        "updated": stats["updated"],
    }
